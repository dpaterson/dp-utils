#!/usr/local/bin/python2.7
# encoding: utf-8
'''
tempest_results_processor -- shortdesc

Takes one or two xml subunit results and generates an xls
report that lists tests unique to each run and compares
the tests common to both runs.

If only a single result is supplied a simple xls report is
generated with the contents of xml file.

@author:     David Paterson

@copyright:  2017 Dell Inc. All rights reserved.
'''

from openpyxl import Workbook
import os
import sys
import time
import xml.dom.minidom as minidom

from argparse import ArgumentParser
from argparse import RawDescriptionHelpFormatter
from collections import namedtuple

Result = namedtuple('Result',
                    'name, path, errors, failures, num_tests, time, tests')

STATUS_ERROR = 'error'
STATUS_FAILURE = 'failure'
STATUS_OK = 'ok'
STATUS_SKIPPED = 'skipped'


class Results(object):

    def __init__(self, args):
        self.args = args
        self.result_1 = self._get_result(args.result_1)
        if args.result_2:
            self.result_2 = self._get_result(args.result_2)
            (self.unique_1, self.unique_2, self.matches) = \
                self._compare(self.result_1.tests, self.result_2.tests)

    def _get_result(self, results_file):
        """Create Result namedtuple from file path
        """
        short_name = os.path.basename(results_file)
        dom = minidom.parse(results_file)
        suite = dom.documentElement
        errors = suite.getAttribute("errors")
        failures = suite.getAttribute("failures")
        num_tests = suite.getAttribute("tests")
        time = suite.getAttribute("time")
        test_cases = suite.getElementsByTagName("testcase")
        test_map = self._list_to_dict(test_cases)

        result = Result(short_name, results_file, errors,
                        failures, num_tests,
                        time, test_map)
        return result

    def _compare(self, tc1, tc2):
        """Compare two sets of tempest test cases.

        :param tc1: first verification test cases json
        :param tc2: second verification test cases json
        """
        names1 = set(tc1.keys())
        names2 = set(tc2.keys())
        match_tests = list(names1.intersection(names2))
        unique_1 = {}
        unique_2 = {}
        matches = {}
        unique_names1 = list(names1.difference(names2))
        unique_names2 = list(names2.difference(names1))
        for name in unique_names1:
            unique_1[name] = tc1[name]
        for name in unique_names2:
            unique_2[name] = tc2[name]
        for name in match_tests:
            matches[name] = tc1[name], tc2[name]

        return unique_1, unique_2, matches

    def _compare_to_xls(self):
        wb = Workbook()

        '''grab the active worksheet'''
        ws_common = wb.active

        ws_common.title = "Matching Test Results"

        '''Create worksheets for each set of unique results'''
        ws_1 = wb.create_sheet(title=self.result_1.name)
        ws_2 = wb.create_sheet(title=self.result_2.name)

        ws_common.append(["File Name", "# Tests", "Errors",
                          "Failures", "Time"])
        ws_common.append([self.result_1.path, self.result_1.num_tests,
                          self.result_1.errors, self.result_1.failures,
                          self.result_1.time])
        ws_common.append([self.result_2.path, self.result_2.num_tests,
                          self.result_2.errors, self.result_2.failures,
                          self.result_2.time])
        ws_common.append([])
        self._add_unique(ws_1, self.result_1.path, self.unique_1)
        self._add_unique(ws_2, self.result_2.path, self.unique_2)
        self._add_common(ws_common, self.result_1.name, self.result_2.name,
                         self.matches)
        '''TODO: improve filename, perhaps use pretty timestamp over
        just using time.time()'''
        output_fn = (self.result_1.name + "-vs-" + self.result_2.name +
                     "_" + str(time.time()) + ".xlsx")
        '''
        TODO play with styles
        c = ws['A1']
        c.font = Font(bold=18, color=colors.RED)

        c.fill = PatternFill(fill_type="solid", bgColor="CFCFCF")
                         # start_color='FFFFFFFF',
                         # end_color='FF000000')
        '''
        wb.save(output_fn)

    def _report_to_xls(self):
        wb = Workbook()

        '''grab the active worksheet'''
        ws = wb.active

        ws.title = "Test Results"

        ws.append(["File Name", "# Tests", "Errors",
                          "Failures", "Time"])
        ws.append([self.result_1.path, self.result_1.num_tests,
                          self.result_1.errors, self.result_1.failures,
                          self.result_1.time])
        self._add_unique(ws, self.result_1.path, self.result_1.tests)

        '''TODO: improve filename, perhaps use pretty timestamp over
        just using time.time()'''
        output_fn = (self.result_1.name + str(time.time()) + ".xlsx")

        wb.save(output_fn)

    def to_xls(self, _is_compare=False):
        """Utilize openpyxl library and output xls results
        TODO: Add some frozen panes and styling to improve UX
        """
        
        if _is_compare:
            self._compare_to_xls()
        else:
            self._report_to_xls()

    def _add_unique(self, worksheet, name, tests):
        worksheet.append(["Number of unique tests in %s: %s" %
                          (name, len(tests))])
        worksheet.append(["Name", "Time", "Status", "Reason"])
        # sort to list of tuples
        tup_list = sorted(tests.items(), key=lambda elem: (elem[1]['status'],
                          elem[0]), reverse=False)

        for test_name, test in tup_list:
            testcase = test["testcase"]
            status = test["status"]
            time = testcase.getAttribute("time")
            reason = ""
            if (status is STATUS_SKIPPED):
                skip_nodes = testcase.getElementsByTagName(STATUS_SKIPPED)
                reason = skip_nodes[0].childNodes[0].data
            elif (status is STATUS_FAILURE):
                failure_nodes = testcase.getElementsByTagName(STATUS_FAILURE)
                reason = failure_nodes[0].childNodes[0].data
                reason = self._trim_failure(reason)
            worksheet.append([test_name, time, status, reason])
        worksheet.append([])

    def _add_common(self, worksheet, name_1, name_2, matches):
        worksheet.append(["Number of common tests: %s" % len(matches)])
        worksheet.append(["Name", "%s-Time" % name_1, "%s-Time" % name_2,
                          "%s-Status" % name_1, "%s-Status" % name_2,
                          "%s-Reason" % name_1, "%s-Reason" % name_2])
        tup_list = sorted(matches.items(), key=lambda elem: (elem[1][0]['status'],
                          elem[0]), reverse=False)

        for match_name, match_tup in tup_list:
            match_1 = match_tup[0]
            match_2 = match_tup[1]
            testcase_1 = match_1["testcase"]
            testcase_2 = match_2["testcase"]
            time_1 = testcase_1.getAttribute("time")
            time_2 = testcase_2.getAttribute("time")
            status_1 = match_1["status"]
            status_2 = match_2["status"]
            reason_1 = ""
            reason_2 = ""

            if (status_1 is STATUS_SKIPPED):
                skip_nodes_1 = testcase_1.getElementsByTagName(STATUS_SKIPPED)
                reason_1 = skip_nodes_1[0].childNodes[0].data
            elif (status_1 is STATUS_FAILURE):
                fail_nodes_1 = testcase_1.getElementsByTagName(STATUS_FAILURE)
                reason_1 = fail_nodes_1[0].childNodes[0].data
                reason_1 = self._trim_failure(reason_1)

            if (status_2 is STATUS_SKIPPED):
                skip_nodes_2 = testcase_2.getElementsByTagName(STATUS_SKIPPED)
                reason_2 = skip_nodes_2[0].childNodes[0].data
            elif (status_2 is STATUS_FAILURE):
                fail_nodes_2 = testcase_2.getElementsByTagName(STATUS_FAILURE)
                reason_2 = fail_nodes_2[0].childNodes[0].data
                reason_2 = self._trim_failure(reason_2)

            worksheet.append([match_name, time_1, time_2, status_1, status_2,
                              reason_1, reason_2])

    def _trim_failure(self, failure):
        indx = failure.rfind("}}}")
        trimmed = ""
        if indx is -1:
            trimmed = (failure[:-2])
        else:
            trimmed = failure[indx+5:-2]

        return trimmed

    def _list_to_dict(self, test_list):

        test_dict = {}
        i = 0
        for testcase in test_list:
            status = STATUS_OK
            if len(testcase.getElementsByTagName(STATUS_SKIPPED)) > 0:
                status = STATUS_SKIPPED
            elif len(testcase.getElementsByTagName(STATUS_FAILURE)) > 0:
                status = STATUS_FAILURE

            class_name = testcase.getAttribute("classname")         #
            name = (class_name + "." + testcase.getAttribute("name")
                    if len(class_name) > 0 else testcase.getAttribute("name"))
            test_dict[name] = {"status": status,
                               "testcase": testcase}
        return test_dict


def main():
        parser = _create_parser()
        args = parser.parse_args()
        results = Results(args)
        results.to_xls((args.result_2 is not None))


def _create_parser():
    parser = ArgumentParser(description="Compare two tempest xml results",
                            formatter_class=RawDescriptionHelpFormatter)
    parser.add_argument("result_1",
                        help="path to xml result 1")
    parser.add_argument("result_2", nargs='?',
                        help="path to xml result 2")
    '''TODO future functionality
    parser.add_argument("-c", "--csv", dest="output_csv",
                        action="store_true",
                        help="output csv")
    parser.add_argument("-m", "--html", dest="output_html",
                        action="store_true",
                        help="output html")
    parser.add_argument("-n", "--json", dest="output_json",
                        action="store_true",
                        help="output json")
    parser.add_argument("-o", "--output-file", dest="output_file",
                        type=str, required=False,
                        help="If specified, output will be saved to given "
                        "file")
    '''
    return parser

if __name__ == "__main__":
    sys.exit(main())
